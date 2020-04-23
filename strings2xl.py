#!/usr/bin/python
# coding: utf-8

# 用于 Android 中 strings.xml 的文本提取生成表格

import xml.sax
import sys
import os

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


xml_name = "strings.xml"
out_file = "strings.xlsx"
stringBeanArr = []


class StringBean():
    def __init__(self, eType, name, content = "", translated = ""):
        self.eType = eType
        self.name = name
        self.content = content
        self.translated = translated


# 解析 strings.xml 的具体实现
class StringsHandler(xml.sax.ContentHandler):
    def __init__(self):
        self.tag = ""
        self.child_tag = ""
        self.key = ""
        self.child_key = ""
        self.value = ""

    def startElement(self, tag, attributes):
        if tag == "resources":
            # resources 不记录
            self.tag = ""
            return
        if tag == "string":
            self.tag = tag
            self.key = attributes["name"]
        elif tag == "plurals":
            self.tag = tag
            self.key = attributes["name"]
        elif tag == "item":
            self.child_tag = tag
            self.child_key = attributes["quantity"]

    def characters(self, content):
        if self.tag == "" or content == "\n" or content == "    "  or content == "        " or content == "\t" or content == "\r":
            return
        self.value += content

    def endElement(self, tag):
        # 将元素实例添加到数组中
        if tag == "resources":
            self.value = ""
            return
        if tag == "item":
            stringBeanArr.append(
                StringBean(self.child_tag, self.child_key, self.value))
        else:
            stringBeanArr.append(StringBean(self.tag, self.key, self.value))
        self.value = ""


# 解析 strings.xml 文件
def parseXML(file_name):
    parser = xml.sax.make_parser()
    parser.setFeature(xml.sax.handler.feature_namespaces, 0)

    # 重写 ContextHandler
    Handler = StringsHandler()
    parser.setContentHandler(Handler)

    parser.parse(file_name)


def cellStringBean(ws: Worksheet, rowIndex, strBean: StringBean):
    ws.cell(row = rowIndex, column = 1, value = strBean.eType)
    ws.cell(row = rowIndex, column = 2, value = strBean.name)
    ws.cell(row = rowIndex, column = 3, value = strBean.content)
    ws.cell(row = rowIndex, column = 4, value = strBean.translated)


def setSheetParams(ws: Worksheet):
    ws.title = xml_name
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60


def main():
    # 解析文件，并将数据实例放到数组中
    parseXML(xml_name)
    # 建立 book
    wb = Workbook()
    # 初始化表
    ws = wb.active
    setSheetParams(ws)

    # 写数据
    i = 1
    cellStringBean(ws, i, StringBean("Type", "Name", "Content", "Translated"))

    for data in stringBeanArr:
        i += 1
        cellStringBean(ws, i, data)

    # 输出表格
    wb.save(out_file)
    

main()
